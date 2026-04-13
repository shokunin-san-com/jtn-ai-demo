"""Excel 書き込みモジュールのスモークテスト"""

from io import BytesIO

from openpyxl import load_workbook

from lib.demo1_kyk.excel_writer import write_kyk
from lib.demo2_sekou_keikaku.excel_writer import write_chapter
from lib.demo3_geppo.excel_writer import write_geppo


MOCK_KYK = {
    "作業内容": ["①作業A", "②作業B", "③作業C", "④作業D"],
    "リスク": [
        {"危険": "感電", "可能性": "△", "重大性": "×", "危険度": 4, "対策": "検電確認"},
    ],
    "重点対策": "感電防止徹底",
    "安全指示": "WBGT確認",
}

MOCK_GEPPO = {
    "工事報告": "報告本文",
    "作業概要": "概要",
    "出来高": {"当月": 10, "累計": 50},
}


def test_write_kyk_produces_valid_xlsx():
    data = write_kyk(MOCK_KYK, date="2026-04-13", weather="晴れ")
    assert len(data) > 1000
    wb = load_workbook(BytesIO(data))
    assert "作業日報・KYK" in wb.sheetnames


def test_write_chapter_all_keys():
    for key in ("ch1", "ch2", "ch7", "ch8"):
        data = write_chapter(key, "テスト本文" * 20)
        assert len(data) > 1000
        wb = load_workbook(BytesIO(data))
        assert len(wb.sheetnames) >= 1


def test_write_geppo_produces_valid_xlsx():
    data = write_geppo(MOCK_GEPPO, target_month="2026年4月", worker_count=5)
    assert len(data) > 1000
    wb = load_workbook(BytesIO(data))
    assert "工事月報" in wb.sheetnames
