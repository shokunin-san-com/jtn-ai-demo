"""工事月報 Excel 書き込みモジュール

templates/demo3/工事月報（月）.xlsx に AI 生成結果を流し込む。
"""

import io
import pathlib

from openpyxl import load_workbook


def _safe_set(ws, coord: str, value) -> None:
    """結合セルの先頭セルを解決して値をセットする。"""
    cell = ws[coord]
    if cell.__class__.__name__ == "MergedCell":
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                ws.cell(row=rng.min_row, column=rng.min_col).value = value
                return
        return
    cell.value = value


TEMPLATE_PATH = (
    pathlib.Path(__file__).resolve().parent.parent.parent
    / "templates"
    / "demo3"
    / "工事月報（月）.xlsx"
)


def write_geppo(ai_data: dict, target_month: str = "", worker_count: int = 0) -> bytes:
    """AI 生成データを月報テンプレートに書き込み、xlsx のバイト列を返す。"""
    wb = load_workbook(TEMPLATE_PATH)

    # 1枚目「工事月報」: ヘッダ情報
    if "工事月報" in wb.sheetnames:
        cover = wb["工事月報"]
        if target_month:
            _safe_set(cover, "O17", f"（{target_month}分）")

    # 2枚目「その２」: 工事進捗状況などの記述部
    if "その２" in wb.sheetnames:
        ws2 = wb["その２"]
        _safe_set(ws2, "A3", ai_data.get("工事報告", ""))
        _safe_set(ws2, "A10", ai_data.get("作業概要", ""))

    # 3枚目「その３」: 出来高数値
    if "その３）" in wb.sheetnames:
        ws3 = wb["その３）"]
        dekidaka = ai_data.get("出来高", {})
        _safe_set(ws3, "AJ2", dekidaka.get("累計", ""))
        if worker_count:
            _safe_set(ws3, "AJ3", worker_count)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
