"""過去案件テキスト抽出モジュール

templates/reference/sekou/ 配下の Excel ファイルからテキストを抽出し、
章別プロンプトの参考資料として利用できる形で返す。
"""

import pathlib
from typing import Dict, List

from openpyxl import load_workbook


REFERENCE_DIR = (
    pathlib.Path(__file__).resolve().parent.parent.parent
    / "templates"
    / "reference"
    / "sekou"
)

# 章キー → 参考ファイル名（存在するファイルのみ対応）
CHAPTER_REFERENCE_MAP = {
    "ch1": "1.総則.xlsx",
    "ch2": "2.工事概要.xlsx",
    "ch7": "7.品質管理.xlsx",
    "ch8": "8.安全衛生管理.xlsx",
}


def extract_text_from_xlsx(path: pathlib.Path, max_chars: int = 3000) -> str:
    """Excel ファイルから全シートの文字列セル値を抽出し、改行で連結する。"""
    if not path.exists():
        return ""
    wb = load_workbook(path, read_only=True, data_only=True)
    parts: List[str] = []
    total = 0
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            for val in row:
                if val is None:
                    continue
                s = str(val).strip()
                if not s:
                    continue
                parts.append(s)
                total += len(s)
                if total >= max_chars:
                    return "\n".join(parts)[:max_chars]
    return "\n".join(parts)


def load_chapter_reference(chapter_key: str, max_chars: int = 3000) -> str:
    """指定章の参考テキストを返す。ファイルが無ければ空文字列。"""
    fname = CHAPTER_REFERENCE_MAP.get(chapter_key)
    if not fname:
        return ""
    return extract_text_from_xlsx(REFERENCE_DIR / fname, max_chars=max_chars)


def load_all_references(max_chars: int = 3000) -> Dict[str, str]:
    """全章分の参考テキストをまとめて返す。"""
    return {
        key: load_chapter_reference(key, max_chars=max_chars)
        for key in CHAPTER_REFERENCE_MAP
    }
