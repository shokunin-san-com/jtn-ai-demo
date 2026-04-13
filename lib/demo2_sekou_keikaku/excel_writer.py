"""施工計画書 Excel 書き込みモジュール

templates/demo2/ の各章テンプレートに AI 生成本文を流し込む。
"""

import io
import pathlib

from openpyxl import load_workbook


TEMPLATE_DIR = (
    pathlib.Path(__file__).resolve().parent.parent.parent / "templates" / "demo2"
)

# 章キー → テンプレートファイル名
CHAPTER_FILES = {
    "ch1": "1.総則.xlsx",
    "ch2": "2.工事概要.xlsx",
    "ch7": "7.品質管理.xlsx",
    "ch8": "8.安全衛生管理.xlsx",
}

# 各章の AI 本文を書き込む位置（本文セクション冒頭の空き領域）
# 既存テンプレートの B6 以降に AI 生成テキストを追加する形式
AI_CONTENT_CELL = "B35"  # シート末尾付近、重複を避けた位置


def write_chapter(chapter_key: str, content: str) -> bytes:
    """指定章の Excel に AI 本文を書き込み、xlsx のバイト列を返す。"""
    if chapter_key not in CHAPTER_FILES:
        raise ValueError(f"未知の章キー: {chapter_key}")

    template_path = TEMPLATE_DIR / CHAPTER_FILES[chapter_key]
    wb = load_workbook(template_path)

    # 本文シート（「目次」「表紙」を除いた最初のシート）を選択
    target_sheet = None
    for sn in wb.sheetnames:
        if "目次" not in sn and "表紙" not in sn:
            target_sheet = wb[sn]
            break
    if target_sheet is None:
        target_sheet = wb.active

    # AI 生成本文を末尾付近に書き込む（結合セル対応）
    cell = target_sheet[AI_CONTENT_CELL]
    value = "【AI生成】\n" + content
    if cell.__class__.__name__ == "MergedCell":
        for rng in target_sheet.merged_cells.ranges:
            if AI_CONTENT_CELL in rng:
                target_sheet.cell(row=rng.min_row, column=rng.min_col).value = value
                break
    else:
        cell.value = value

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
