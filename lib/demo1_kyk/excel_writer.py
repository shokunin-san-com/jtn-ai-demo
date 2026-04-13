"""KYKシート Excel 書き込みモジュール

テンプレート (templates/demo1/KYKシート.xlsx) に AI 生成結果を流し込む。
"""

import io
import pathlib

from openpyxl import load_workbook


TEMPLATE_PATH = (
    pathlib.Path(__file__).resolve().parent.parent.parent
    / "templates"
    / "demo1"
    / "KYKシート.xlsx"
)

# 作業内容を書き込むセル (①②③④)
WORK_CELLS = ["D5", "D7", "D9", "D11"]

# リスク行の開始位置。最大3行（行14,16,18）
RISK_ROWS = [14, 16, 18]
# 各行の列マッピング
RISK_COLUMNS = {
    "危険": "B",
    "可能性": "G",
    "重大性": "H",
    "評価": "I",
    "危険度": "J",
    "対策": "K",
}


def _safe_set(ws, coord: str, value) -> None:
    """結合セルの先頭セルを解決して値をセットする。"""
    cell = ws[coord]
    # MergedCell の場合、包含する結合範囲の左上セルに書き込む
    if cell.__class__.__name__ == "MergedCell":
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                ws.cell(row=rng.min_row, column=rng.min_col).value = value
                return
        return
    cell.value = value


def _evaluation_symbol(possibility: str, severity: str) -> str:
    """可能性と重大性から評価記号を組み立てる（例: △×）"""
    return f"{possibility or ''}{severity or ''}"


def write_kyk(ai_data: dict, date: str = "", weather: str = "") -> bytes:
    """AI生成データを KYK テンプレートに書き込み、xlsx のバイト列を返す。"""
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active  # 「作業日報・KYK」シート

    # 日付（C2 は結合セルなので安全に設定）
    if date:
        _safe_set(ws, "C2", date)
    # 天候は K2 セルにラベルがあるので、L2 へ
    if weather:
        _safe_set(ws, "L2", weather)

    # 作業内容
    items = ai_data.get("作業内容", [])
    for cell, item in zip(WORK_CELLS, items):
        _safe_set(ws, cell, item)

    # リスク
    risks = ai_data.get("リスク", [])[: len(RISK_ROWS)]
    for row, risk in zip(RISK_ROWS, risks):
        _safe_set(ws, f"{RISK_COLUMNS['危険']}{row}", risk.get("危険", ""))
        _safe_set(ws, f"{RISK_COLUMNS['可能性']}{row}", risk.get("可能性", ""))
        _safe_set(ws, f"{RISK_COLUMNS['重大性']}{row}", risk.get("重大性", ""))
        _safe_set(
            ws,
            f"{RISK_COLUMNS['評価']}{row}",
            _evaluation_symbol(risk.get("可能性", ""), risk.get("重大性", "")),
        )
        _safe_set(ws, f"{RISK_COLUMNS['危険度']}{row}", risk.get("危険度", ""))
        _safe_set(ws, f"{RISK_COLUMNS['対策']}{row}", risk.get("対策", ""))

    # 重点対策・安全指示
    if "重点対策" in ai_data:
        _safe_set(ws, "A21", ai_data["重点対策"])
    if "安全指示" in ai_data:
        _safe_set(ws, "A23", ai_data["安全指示"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
